#!/usr/bin/env python3
"""
Build a queryable journal-metrics SQLite DB from a local JCR+CAS Excel file.

This script is meant to be run once (or when the Excel source updates) and the
resulting DB should live inside this skill folder so other scripts can query it.

Default source discovery (Windows):
  - Look for the first *.xlsx file under D:/game whose name contains "JCR".

Output:
  - ../references/journal_metrics_2025.sqlite3

Dependencies:
  - openpyxl (already present in most Anaconda installs)
"""

from __future__ import annotations

import argparse
import os
import re
import sqlite3
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Iterator

import openpyxl


def normalize_journal_name(value: str) -> str:
    value = unicodedata.normalize("NFKC", str(value)).strip().lower()
    return "".join(ch for ch in value if ch.isalnum())


def normalize_issn(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip().upper()
    text = re.sub(r"[^0-9X]", "", text)
    return text or None


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def to_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(float(str(value).strip()))
    except ValueError:
        return None


@dataclass(frozen=True)
class JcrRow:
    journal: str
    journal_norm: str
    jif_2024: float | None
    quartile: str | None
    jif_rank: str | None
    partition_2023: str | None
    total_citation: int | None
    category: str | None
    issn: str | None
    issn_norm: str | None
    eissn: str | None
    eissn_norm: str | None


@dataclass(frozen=True)
class CasRow:
    journal: str
    journal_norm: str
    partition_2025: int | None
    top: str | None
    open_access: str | None


def iter_rows(ws: Any, min_row: int) -> Iterator[tuple[Any, ...]]:
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        yield row


def find_default_xlsx() -> Path:
    base = Path("D:/game")
    if not base.exists():
        raise FileNotFoundError("Default search directory not found: D:/game")
    candidates = [
        base / name
        for name in os.listdir(base)
        if name.lower().endswith(".xlsx") and "JCR" in name
    ]
    if not candidates:
        raise FileNotFoundError('No *.xlsx with "JCR" found under D:/game')
    candidates.sort(key=lambda p: p.name)
    return candidates[0]


def detect_sheets(wb: Any) -> tuple[Any, Any]:
    jcr_sheet_names = [s for s in wb.sheetnames if "JCR" in s.upper() or "JIF" in s.upper()]
    if not jcr_sheet_names:
        raise ValueError("Unable to detect the JCR sheet (no sheet name contains JCR/JIF).")
    jcr_ws = wb[jcr_sheet_names[0]]

    cas_sheet_names = [s for s in wb.sheetnames if s != jcr_ws.title]
    if not cas_sheet_names:
        raise ValueError("Unable to detect the CAS sheet (only one sheet found).")
    cas_ws = wb[cas_sheet_names[0]]
    return jcr_ws, cas_ws


def parse_jcr(ws: Any) -> list[JcrRow]:
    rows: list[JcrRow] = []
    # Layout (based on your file): row 2 = headers, row 3+ = data
    for r in iter_rows(ws, min_row=3):
        journal = r[0] if len(r) > 0 else None
        if journal is None or str(journal).strip() == "":
            continue
        journal_str = str(journal).strip()
        rows.append(
            JcrRow(
                journal=journal_str,
                journal_norm=normalize_journal_name(journal_str),
                jif_2024=to_float(r[1] if len(r) > 1 else None),
                quartile=str(r[2]).strip() if len(r) > 2 and r[2] is not None else None,
                jif_rank=str(r[3]).strip() if len(r) > 3 and r[3] is not None else None,
                partition_2023=str(r[4]).strip() if len(r) > 4 and r[4] is not None else None,
                total_citation=to_int(r[5] if len(r) > 5 else None),
                category=str(r[6]).strip() if len(r) > 6 and r[6] is not None else None,
                issn=str(r[7]).strip() if len(r) > 7 and r[7] is not None else None,
                issn_norm=normalize_issn(r[7] if len(r) > 7 else None),
                eissn=str(r[8]).strip() if len(r) > 8 and r[8] is not None else None,
                eissn_norm=normalize_issn(r[8] if len(r) > 8 else None),
            )
        )
    return rows


def parse_cas(ws: Any) -> list[CasRow]:
    rows: list[CasRow] = []
    for r in iter_rows(ws, min_row=3):
        journal = r[0] if len(r) > 0 else None
        if journal is None or str(journal).strip() == "":
            continue
        journal_str = str(journal).strip()
        rows.append(
            CasRow(
                journal=journal_str,
                journal_norm=normalize_journal_name(journal_str),
                partition_2025=to_int(r[1] if len(r) > 1 else None),
                top=str(r[2]).strip() if len(r) > 2 and r[2] is not None else None,
                open_access=str(r[3]).strip() if len(r) > 3 and r[3] is not None else None,
            )
        )
    return rows


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        PRAGMA journal_mode=WAL;
        PRAGMA synchronous=NORMAL;

        DROP TABLE IF EXISTS jcr;
        DROP TABLE IF EXISTS cas;

        CREATE TABLE jcr (
          id INTEGER PRIMARY KEY,
          journal TEXT NOT NULL,
          journal_norm TEXT NOT NULL,
          jif_2024 REAL,
          quartile TEXT,
          jif_rank TEXT,
          partition_2023 TEXT,
          total_citation INTEGER,
          category TEXT,
          issn TEXT,
          issn_norm TEXT,
          eissn TEXT,
          eissn_norm TEXT
        );

        CREATE INDEX idx_jcr_journal_norm ON jcr(journal_norm);
        CREATE INDEX idx_jcr_issn_norm ON jcr(issn_norm);
        CREATE INDEX idx_jcr_eissn_norm ON jcr(eissn_norm);

        CREATE TABLE cas (
          id INTEGER PRIMARY KEY,
          journal TEXT NOT NULL,
          journal_norm TEXT NOT NULL,
          partition_2025 INTEGER,
          top TEXT,
          open_access TEXT
        );

        CREATE INDEX idx_cas_journal_norm ON cas(journal_norm);
        """
    )


def insert_rows(conn: sqlite3.Connection, jcr_rows: Iterable[JcrRow], cas_rows: Iterable[CasRow]) -> None:
    conn.executemany(
        """
        INSERT INTO jcr(
          journal, journal_norm, jif_2024, quartile, jif_rank, partition_2023,
          total_citation, category, issn, issn_norm, eissn, eissn_norm
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                r.journal,
                r.journal_norm,
                r.jif_2024,
                r.quartile,
                r.jif_rank,
                r.partition_2023,
                r.total_citation,
                r.category,
                r.issn,
                r.issn_norm,
                r.eissn,
                r.eissn_norm,
            )
            for r in jcr_rows
        ],
    )
    conn.executemany(
        """
        INSERT INTO cas(
          journal, journal_norm, partition_2025, top, open_access
        ) VALUES (?, ?, ?, ?, ?)
        """,
        [
            (
                r.journal,
                r.journal_norm,
                r.partition_2025,
                r.top,
                r.open_access,
            )
            for r in cas_rows
        ],
    )


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build journal_metrics_2025.sqlite3 from an Excel workbook.")
    parser.add_argument(
        "--xlsx",
        type=Path,
        help="Path to the source .xlsx file. If omitted, auto-discover under D:/game.",
    )
    parser.add_argument(
        "--out",
        type=Path,
        help="Output DB path (default: ../references/journal_metrics_2025.sqlite3).",
    )
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing output DB.")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    xlsx_path: Path = args.xlsx or find_default_xlsx()
    out_path: Path = args.out or (Path(__file__).resolve().parents[1] / "references" / "journal_metrics_2025.sqlite3")

    if not xlsx_path.is_file():
        raise SystemExit(f"Excel file not found: {xlsx_path}")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    if out_path.exists() and not args.overwrite:
        raise SystemExit(f"Output DB already exists: {out_path} (use --overwrite to replace)")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    jcr_ws, cas_ws = detect_sheets(wb)

    jcr_rows = parse_jcr(jcr_ws)
    cas_rows = parse_cas(cas_ws)

    if out_path.exists():
        out_path.unlink()
    conn = sqlite3.connect(out_path)
    try:
        create_schema(conn)
        insert_rows(conn, jcr_rows, cas_rows)
        conn.commit()
    finally:
        conn.close()

    print(f"Built DB: {out_path}")
    print(f"Rows: jcr={len(jcr_rows)} cas={len(cas_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

